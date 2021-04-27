#!/usr/bin/env python
# coding: utf-8

# In[51]:


import pandas as pd 
import requests 
import csv

df = pd.read_excel('한국어어휘사전_표제어_주제분류_20210210.xlsx', sheet_name='표제어 목록',engine="openpyxl")
tempList=dict()
for idx in range(len(df)):
    temp=reversed(list(df['f'][idx]))
    string=''
    tempList[df['a'][idx]]=df['f'][idx]
    if temp is not None:
        for s in temp:
            string=string+s
        string=re.findall('[^\d^/^ ^∙^\.]+',string)[0]
        df.loc[idx,'f']= string
import re
df=df.sort_values(by='f',ascending=False)
for idx in range(len(df)): 
    df.loc[idx,'f']= tempList[df['a'][idx]]

df.to_csv('정렬역순.csv',encoding='utf-8-sig')


# In[47]:





# In[3]:


from openpyxl import load_workbook


# In[4]:


load_wb = load_workbook("한국어어휘사전_표제어_주제분류_20210210.xlsx", data_only=True)
load_ws = load_wb['표제어 목록']


# In[23]:


a=load_ws['B3']

a.fill
for row in load_ws.rows:
    print(row[0].value)


# In[39]:


import pandas as pd 
import requests 
import csv
from openpyxl import load_workbook
import re
load_wb = load_workbook("세종_한국어어휘사전_표제어_용례_20201117.xlsx", data_only=True)
load_ws = load_wb['표제어 목록']
tempList=dict()
for row in load_ws.rows:
    temp=reversed(list(row[6].value))
    string=''
    tempList[row[0].value]=row[6].value
    if temp is not None:
        for s in temp:
            string=string+s
        string=re.findall('[^\d^/^ ^∙^\.]+',string)[0]
        row[6].value= string


load_wb.save('test.xlsx')



# In[41]:



load_wb1 = load_workbook("test.xlsx", data_only=True)
load_ws1 = load_wb1['표제어 목록']
print(load_ws1)
for row in load_ws1.rows:
    print([row[6].value,tempList[row[0].value]])
    row[6].value=tempList[row[0].value]
load_wb1.save('세종_한국어어휘사전_표제어_용례_20201117_역순정렬.xlsx')


# In[27]:


print(lo)

